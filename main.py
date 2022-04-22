# -*- coding: utf-8 -*-
import time
import psutil
import logging
import threading
from datetime import datetime
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook
from pynput.mouse import Controller
from pynput import keyboard


def check_device():
    """Проверяет общую активность устройств в случае
    активности выводит True,
    если девайсы не используются выводит False """
    def check_py():
        """Проверяем активность процесса пйчарма"""
        for proc in psutil.process_iter():
            if proc.name() == 'pycharm.sh':
                return True
            else:
                continue

    def check_mouse():
        """Проверяет активность мыши и возвращает
        True или False"""
        logging.debug('check device', )
        mouse = Controller()
        check = mouse.position
        time.sleep(20)
        if mouse.position != check:
            return True
        else:
            return False

    def check_keyBoard():
        """Проверяет активность клавиатуры
        и возвращает True или False"""
        with keyboard.Events() as events:
            event = events.get(5.0)
            if event is None:
                return False
            else:
                return True

    if check_mouse() and check_py():
        return True
    elif check_keyBoard() and check_py():
        return True
    else:
        return False


def check_video():
    """Проверяем активность процесса видеоплейера"""
    try:
        for proc in psutil.process_iter():
            if proc.name() == 'totem':
                for totem in proc.as_dict()['open_files']:
                    path = totem.path
                    if path.startswith('/home/ivan/Видео/Лекции') \
                            and proc.cpu_percent(interval=1) > 0:
                        return True
        return False
    except Exception:
        return False


def get_time_proc(lock):
    """Записывает историю по работе юзера в приложении"""
    while True:
        logging.debug('running')
        time.sleep(5)
        if check_device():
            time_start = datetime.now()
            while True:
                time.sleep(3.25)
                time_end = datetime.now()
                if not check_device():
                    result = time_end - time_start
                    print(f'Работал в пайчарме {result.seconds} секунды')
                    making_note(
                        day=time_end.day,
                        full_data=time_end.date(),
                        time_of_start=time_start.time(),
                        time_of_end=time_end.time(),
                        proc='paycharm',
                        second=result.seconds,
                        lock=lock
                    )
                    break


def get_time_video(lock):
    """Записывает историю при просмотре обучающего
    видео юзером"""
    while True:
        time.sleep(5)
        logging.debug('running')
        if check_video():
            time_start = datetime.now()
            while True:
                time.sleep(2)
                if not check_video():
                    time_end = datetime.now()
                    result = time_end - time_start
                    print(f'Смотрел обучающий видос {result.seconds} секунд')
                    making_note(
                        day=time_end.day,
                        full_data=time_end.date(),
                        time_of_start=time_start.time(),
                        time_of_end=time_end.time(),
                        proc='watch_video',
                        second=result.seconds,
                        lock=lock
                    )
                    break


def making_note(day, full_data, time_of_start, time_of_end, proc, second, lock):
    """Выполняет запись в таблицу XL
    :param: Day
    :param: full_Data
    :param: time_start
    :param: time_end
    :param: second
    """
    lock.acquire()
    filename = 'timer.xlsx'
    font = Font(
        name='Calibri',
        size=20,
        bold=True,
        italic=False,
        strike=False,
        color='FF000000'
    )
    alignment = Alignment(
        horizontal='general',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0
    )
    logging.debug('Send massage')
    wb = load_workbook(filename, read_only=False, keep_links=True, keep_vba=False)
    sheet = str(datetime.now().month)
    try:
        ws = wb[sheet]
        ws.append([day, full_data, time_of_start, time_of_end, proc, second])
    except Exception:
        ws = wb.create_sheet(sheet)
        ws.append(['DAY', 'FULL_DATA', 'TIME_OF_START', 'TIME_OF_END', 'PROC', 'SECOND'])
        ws.append([day, full_data, time_of_start, time_of_end, proc, second])
        columns = ['A', 'B', 'C', 'D', 'E', 'F']
        for column in columns:
            col = ws.column_dimensions[column]
            col.width = 35
            col.alignment = Alignment(horizontal='center')
        row = ws.row_dimensions[1]
        row.font = font
        row.height = 25
        row.alignment = alignment
    finally:
        lock.release()
        wb.save(filename)
        wb.close()


logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s (%(threadName)-10s) %(message)s",
)


def main():
    lock = threading.Lock()
    task_1 = threading.Thread(name='check_video', target=get_time_video, daemon=True, args=(lock,))
    task_2 = threading.Thread(name='get_time_proc', target=get_time_proc, args=(lock,))
    task_1.start()
    task_2.start()


if __name__ == '__main__':
    main()
